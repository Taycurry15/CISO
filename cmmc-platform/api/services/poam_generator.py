"""
POA&M (Plan of Action & Milestones) Generator Service

Generates CMMC Plan of Action & Milestones for remediation tracking
of non-compliant controls.
"""

import logging
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta
from dataclasses import dataclass
from enum import Enum
import asyncpg
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class RiskLevel(str, Enum):
    """Risk level for POA&M items"""
    VERY_HIGH = "Very High"
    HIGH = "High"
    MODERATE = "Moderate"
    LOW = "Low"


class RemediationStatus(str, Enum):
    """Remediation status"""
    OPEN = "Open"
    IN_PROGRESS = "In Progress"
    COMPLETED = "Completed"
    RISK_ACCEPTED = "Risk Accepted"
    DELAYED = "Delayed"


@dataclass
class POAMItem:
    """POA&M remediation item"""
    item_id: str
    control_id: str
    control_title: str
    weakness_description: str
    risk_level: RiskLevel
    impact: str
    likelihood: str
    remediation_plan: str
    resources_required: str
    milestone_date: datetime
    responsible_person: str
    status: RemediationStatus
    completion_date: Optional[datetime]
    cost_estimate: Optional[float]
    comments: Optional[str]


@dataclass
class POAMMetadata:
    """POA&M document metadata"""
    system_name: str
    organization: str
    prepared_by: str
    preparation_date: datetime
    review_date: Optional[datetime]
    version: str


class POAMGenerator:
    """
    POA&M Generator Service

    Generates Plans of Action & Milestones for remediation tracking
    with AI-powered recommendations and risk assessments.
    """

    def __init__(
        self,
        db_pool: asyncpg.Pool,
        ai_service: Optional[Any] = None
    ):
        """
        Initialize POA&M generator

        Args:
            db_pool: Database connection pool
            ai_service: AI service for generating remediation recommendations
        """
        self.db_pool = db_pool
        self.ai_service = ai_service

    async def generate_poam(
        self,
        assessment_id: str,
        metadata: POAMMetadata,
        generate_recommendations: bool = True,
        auto_assign_risk: bool = True
    ) -> BytesIO:
        """
        Generate POA&M document

        Args:
            assessment_id: Assessment ID
            metadata: POA&M metadata
            generate_recommendations: Use AI to generate remediation recommendations
            auto_assign_risk: Automatically calculate risk levels

        Returns:
            BytesIO: Generated Excel workbook
        """
        logger.info(f"Generating POA&M for assessment {assessment_id}")

        # Get non-compliant controls
        items = await self._get_poam_items(
            assessment_id,
            generate_recommendations,
            auto_assign_risk
        )

        # Generate Excel workbook
        wb = self._create_poam_workbook(items, metadata)

        # Save to BytesIO
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        logger.info(f"POA&M generation complete. {len(items)} items included.")
        return excel_bytes

    async def _get_poam_items(
        self,
        assessment_id: str,
        generate_recommendations: bool,
        auto_assign_risk: bool
    ) -> List[POAMItem]:
        """Get POA&M items from assessment findings"""
        items = []

        async with self.db_pool.acquire() as conn:
            # Get "Not Met" and "Partially Met" controls
            findings = await conn.fetch("""
                SELECT
                    cf.id as finding_id,
                    cf.control_id,
                    c.title as control_title,
                    cf.status,
                    cf.assessor_narrative,
                    cf.ai_confidence_score,
                    c.requirement as control_requirement
                FROM control_findings cf
                JOIN controls c ON cf.control_id = c.id
                WHERE cf.assessment_id = $1
                  AND cf.status IN ('Not Met', 'Partially Met')
                ORDER BY cf.control_id
            """, assessment_id)

            for idx, finding in enumerate(findings, 1):
                # Generate weakness description
                weakness = self._generate_weakness_description(finding)

                # Generate remediation plan
                if generate_recommendations and self.ai_service:
                    remediation = await self._generate_remediation_plan(finding)
                else:
                    remediation = "Remediation plan to be developed"

                # Calculate risk level
                if auto_assign_risk:
                    risk_level = self._calculate_risk_level(finding)
                else:
                    risk_level = RiskLevel.MODERATE

                # Estimate milestone date (90 days for High, 180 for Moderate, 365 for Low)
                days_to_remediate = {
                    RiskLevel.VERY_HIGH: 30,
                    RiskLevel.HIGH: 90,
                    RiskLevel.MODERATE: 180,
                    RiskLevel.LOW: 365
                }
                milestone_date = datetime.utcnow() + timedelta(days=days_to_remediate[risk_level])

                # Create POA&M item
                item = POAMItem(
                    item_id=f"POAM-{idx:03d}",
                    control_id=finding['control_id'],
                    control_title=finding['control_title'],
                    weakness_description=weakness,
                    risk_level=risk_level,
                    impact=self._determine_impact(risk_level),
                    likelihood=self._determine_likelihood(finding),
                    remediation_plan=remediation,
                    resources_required="TBD",
                    milestone_date=milestone_date,
                    responsible_person="TBD",
                    status=RemediationStatus.OPEN,
                    completion_date=None,
                    cost_estimate=None,
                    comments=None
                )

                items.append(item)

        return items

    def _generate_weakness_description(self, finding: Dict[str, Any]) -> str:
        """Generate weakness description from finding"""
        if finding['status'] == 'Not Met':
            return f"Control {finding['control_id']} is not implemented. {finding['assessor_narrative'][:200] if finding['assessor_narrative'] else ''}"
        else:  # Partially Met
            return f"Control {finding['control_id']} is partially implemented. {finding['assessor_narrative'][:200] if finding['assessor_narrative'] else ''}"

    async def _generate_remediation_plan(self, finding: Dict[str, Any]) -> str:
        """Generate AI-powered remediation plan"""
        # Placeholder - would use AI service to generate detailed remediation steps
        # For now, return template
        return f"""1. Review current implementation of {finding['control_id']}
2. Identify gaps and deficiencies
3. Develop implementation plan with milestones
4. Implement required controls and procedures
5. Test and validate implementation
6. Document implementation and gather evidence
7. Request re-assessment"""

    def _calculate_risk_level(self, finding: Dict[str, Any]) -> RiskLevel:
        """Calculate risk level based on control and status"""
        # Simple heuristic - would be more sophisticated in production
        # Based on control domain and implementation status

        control_id = finding['control_id']

        # High-risk domains
        high_risk_domains = ['AC', 'IA', 'SC', 'AU']  # Access Control, Identity, Crypto, Audit

        # Extract domain from control ID (e.g., "AC.L2-3.1.1" -> "AC")
        domain = control_id.split('.')[0] if '.' in control_id else ''

        if finding['status'] == 'Not Met':
            if domain in high_risk_domains:
                return RiskLevel.VERY_HIGH
            else:
                return RiskLevel.HIGH
        else:  # Partially Met
            if domain in high_risk_domains:
                return RiskLevel.HIGH
            else:
                return RiskLevel.MODERATE

    def _determine_impact(self, risk_level: RiskLevel) -> str:
        """Determine impact description"""
        impact_map = {
            RiskLevel.VERY_HIGH: "Critical - Severe impact to CUI confidentiality, integrity, or availability",
            RiskLevel.HIGH: "High - Significant impact to CUI protection",
            RiskLevel.MODERATE: "Moderate - Moderate impact to security posture",
            RiskLevel.LOW: "Low - Minor impact to security posture"
        }
        return impact_map[risk_level]

    def _determine_likelihood(self, finding: Dict[str, Any]) -> str:
        """Determine likelihood of exploitation"""
        # Simple heuristic based on confidence score
        confidence = finding.get('ai_confidence_score', 0.5)

        if confidence < 0.5:
            return "High - Low confidence in assessment, likely exploitable"
        elif confidence < 0.75:
            return "Moderate - Moderate confidence, possibly exploitable"
        else:
            return "Low - High confidence assessment, less likely exploitable"

    def _create_poam_workbook(
        self,
        items: List[POAMItem],
        metadata: POAMMetadata
    ) -> Workbook:
        """Create Excel workbook for POA&M"""
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Add POA&M sheet
        self._add_poam_sheet(wb, items, metadata)

        # Add summary sheet
        self._add_summary_sheet(wb, items, metadata)

        # Add instructions sheet
        self._add_instructions_sheet(wb)

        return wb

    def _add_poam_sheet(
        self,
        wb: Workbook,
        items: List[POAMItem],
        metadata: POAMMetadata
    ):
        """Add main POA&M tracking sheet"""
        ws = wb.create_sheet("POA&M Items")

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells('A1:O1')
        title_cell = ws['A1']
        title_cell.value = f"PLAN OF ACTION & MILESTONES (POA&M) - {metadata.system_name}"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        # Metadata
        ws['A2'] = "Organization:"
        ws['B2'] = metadata.organization
        ws['A3'] = "Prepared By:"
        ws['B3'] = metadata.prepared_by
        ws['A4'] = "Date:"
        ws['B4'] = metadata.preparation_date.strftime("%Y-%m-%d")
        ws['A5'] = "Version:"
        ws['B5'] = metadata.version

        # Column headers (row 7)
        headers = [
            "POA&M ID",
            "Control ID",
            "Control Title",
            "Weakness Description",
            "Risk Level",
            "Impact",
            "Likelihood",
            "Remediation Plan",
            "Resources Required",
            "Milestone Date",
            "Responsible Person",
            "Status",
            "Completion Date",
            "Cost Estimate",
            "Comments"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Set column widths
        column_widths = [12, 15, 25, 40, 12, 35, 35, 40, 20, 15, 20, 15, 15, 12, 30]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Data rows
        for row_num, item in enumerate(items, 8):
            ws.cell(row=row_num, column=1, value=item.item_id)
            ws.cell(row=row_num, column=2, value=item.control_id)
            ws.cell(row=row_num, column=3, value=item.control_title)
            ws.cell(row=row_num, column=4, value=item.weakness_description)

            # Risk level with color coding
            risk_cell = ws.cell(row=row_num, column=5, value=item.risk_level.value)
            risk_cell.font = Font(bold=True)
            if item.risk_level == RiskLevel.VERY_HIGH:
                risk_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                risk_cell.font = Font(color="FFFFFF", bold=True)
            elif item.risk_level == RiskLevel.HIGH:
                risk_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            elif item.risk_level == RiskLevel.MODERATE:
                risk_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            ws.cell(row=row_num, column=6, value=item.impact)
            ws.cell(row=row_num, column=7, value=item.likelihood)
            ws.cell(row=row_num, column=8, value=item.remediation_plan)
            ws.cell(row=row_num, column=9, value=item.resources_required)
            ws.cell(row=row_num, column=10, value=item.milestone_date.strftime("%Y-%m-%d"))
            ws.cell(row=row_num, column=11, value=item.responsible_person)

            # Status with color coding
            status_cell = ws.cell(row=row_num, column=12, value=item.status.value)
            if item.status == RemediationStatus.COMPLETED:
                status_cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                status_cell.font = Font(color="FFFFFF", bold=True)
            elif item.status == RemediationStatus.IN_PROGRESS:
                status_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

            ws.cell(row=row_num, column=13, value=item.completion_date.strftime("%Y-%m-%d") if item.completion_date else "")
            ws.cell(row=row_num, column=14, value=item.cost_estimate or "")
            ws.cell(row=row_num, column=15, value=item.comments or "")

            # Apply borders and text wrapping
            for col_num in range(1, 16):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Freeze panes (headers)
        ws.freeze_panes = 'A8'

    def _add_summary_sheet(
        self,
        wb: Workbook,
        items: List[POAMItem],
        metadata: POAMMetadata
    ):
        """Add summary dashboard sheet"""
        ws = wb.create_sheet("Summary", 0)  # Insert as first sheet

        # Title
        ws['A1'] = "POA&M Summary Dashboard"
        ws['A1'].font = Font(size=16, bold=True)

        # Metadata
        ws['A3'] = "System:"
        ws['B3'] = metadata.system_name
        ws['A4'] = "Total Items:"
        ws['B4'] = len(items)
        ws['A5'] = "Report Date:"
        ws['B5'] = metadata.preparation_date.strftime("%Y-%m-%d")

        # By risk level
        ws['A7'] = "Items by Risk Level"
        ws['A7'].font = Font(bold=True, size=12)

        risk_counts = {level: 0 for level in RiskLevel}
        for item in items:
            risk_counts[item.risk_level] += 1

        row = 8
        for risk_level, count in risk_counts.items():
            ws.cell(row=row, column=1, value=risk_level.value)
            ws.cell(row=row, column=2, value=count)
            row += 1

        # By status
        ws['A13'] = "Items by Status"
        ws['A13'].font = Font(bold=True, size=12)

        status_counts = {status: 0 for status in RemediationStatus}
        for item in items:
            status_counts[item.status] += 1

        row = 14
        for status, count in status_counts.items():
            ws.cell(row=row, column=1, value=status.value)
            ws.cell(row=row, column=2, value=count)
            row += 1

        # Overdue items
        ws['A19'] = "Overdue Items"
        ws['A19'].font = Font(bold=True, size=12)

        overdue_count = sum(1 for item in items if item.milestone_date < datetime.utcnow() and item.status != RemediationStatus.COMPLETED)
        ws['B19'] = overdue_count

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

    def _add_instructions_sheet(self, wb: Workbook):
        """Add instructions sheet"""
        ws = wb.create_sheet("Instructions")

        ws['A1'] = "POA&M Instructions"
        ws['A1'].font = Font(size=14, bold=True)

        instructions = [
            "",
            "1. Review each POA&M item in the 'POA&M Items' sheet",
            "",
            "2. Assign a responsible person for each item",
            "",
            "3. Estimate resources required (personnel, tools, budget)",
            "",
            "4. Update the status as work progresses:",
            "   • Open: Not yet started",
            "   • In Progress: Remediation underway",
            "   • Completed: Remediation finished and validated",
            "   • Risk Accepted: Leadership accepts the risk",
            "   • Delayed: Remediation delayed (add reason in comments)",
            "",
            "5. Update completion date when remediation is finished",
            "",
            "6. Review POA&M monthly and update milestone dates as needed",
            "",
            "7. Use the Summary sheet to track overall progress",
            "",
            "Risk Levels:",
            "   • Very High: Remediate within 30 days",
            "   • High: Remediate within 90 days",
            "   • Moderate: Remediate within 180 days",
            "   • Low: Remediate within 365 days"
        ]

        for row, instruction in enumerate(instructions, 2):
            ws.cell(row=row, column=1, value=instruction)

        ws.column_dimensions['A'].width = 80


# Helper functions
async def generate_poam_for_assessment(
    db_pool: asyncpg.Pool,
    assessment_id: str,
    metadata: POAMMetadata,
    ai_service: Optional[Any] = None
) -> BytesIO:
    """
    Convenience function to generate POA&M

    Args:
        db_pool: Database pool
        assessment_id: Assessment ID
        metadata: POA&M metadata
        ai_service: Optional AI service

    Returns:
        BytesIO: Generated Excel workbook
    """
    generator = POAMGenerator(db_pool, ai_service)
    return await generator.generate_poam(assessment_id, metadata)

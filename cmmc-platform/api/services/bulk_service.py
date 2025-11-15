"""
Bulk Operations Service

Provides batch processing capabilities for large-scale operations:
- Bulk control status updates
- Bulk evidence upload (ZIP files)
- Excel import/export for control findings
- Mass control assignments
- Batch AI analysis requests

Includes progress tracking and error handling for all operations.
"""

import logging
import uuid
import zipfile
import io
from typing import List, Dict, Any, Optional, BinaryIO
from datetime import datetime
from enum import Enum

import asyncpg
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class BulkOperationType(Enum):
    """Types of bulk operations"""
    CONTROL_UPDATE = "control_update"
    EVIDENCE_UPLOAD = "evidence_upload"
    EXCEL_IMPORT = "excel_import"
    MASS_ASSIGNMENT = "mass_assignment"
    BATCH_AI_ANALYSIS = "batch_ai_analysis"


class BulkOperationStatus(Enum):
    """Status of bulk operation"""
    PENDING = "pending"
    IN_PROGRESS = "in_progress"
    COMPLETED = "completed"
    FAILED = "failed"
    PARTIAL = "partial"


class ControlStatus(Enum):
    """Control finding status"""
    NOT_STARTED = "Not Started"
    IN_PROGRESS = "In Progress"
    MET = "Met"
    NOT_MET = "Not Met"
    PARTIALLY_MET = "Partially Met"
    NOT_APPLICABLE = "Not Applicable"


class BulkService:
    """
    Service for bulk operations on assessments

    Handles batch processing with progress tracking and error handling.
    """

    def __init__(self, db_pool: asyncpg.Pool):
        self.db_pool = db_pool

    # ========================================================================
    # Bulk Control Updates
    # ========================================================================

    async def bulk_update_control_status(
        self,
        assessment_id: str,
        updates: List[Dict[str, Any]],
        updated_by: str
    ) -> Dict[str, Any]:
        """
        Update status for multiple controls at once

        Args:
            assessment_id: Assessment ID
            updates: List of updates with control_id and status
                [{"control_id": "AC.L2-3.1.1", "status": "Met", ...}, ...]
            updated_by: User ID performing the update

        Returns:
            dict: Results with success/failure counts

        Example:
            updates = [
                {
                    "control_id": "AC.L2-3.1.1",
                    "status": "Met",
                    "implementation_narrative": "Access control implemented",
                    "risk_level": "Low"
                },
                {
                    "control_id": "AC.L2-3.1.2",
                    "status": "Partially Met",
                    "findings": "Some gaps identified"
                }
            ]
            result = await service.bulk_update_control_status(
                assessment_id, updates, user_id
            )
        """
        success_count = 0
        failure_count = 0
        errors = []

        async with self.db_pool.acquire() as conn:
            async with conn.transaction():
                for update in updates:
                    try:
                        control_id = update.get('control_id')
                        status = update.get('status')

                        if not control_id or not status:
                            errors.append({
                                'control_id': control_id,
                                'error': 'Missing control_id or status'
                            })
                            failure_count += 1
                            continue

                        # Build update query dynamically
                        set_clauses = ['status = $3', 'updated_at = NOW()', 'created_by = $4']
                        params = [assessment_id, control_id, status, updated_by]
                        param_index = 5

                        # Optional fields
                        optional_fields = [
                            'implementation_status', 'implementation_narrative',
                            'test_results', 'findings', 'recommendations',
                            'risk_level', 'residual_risk'
                        ]

                        for field in optional_fields:
                            if field in update:
                                set_clauses.append(f'{field} = ${param_index}')
                                params.append(update[field])
                                param_index += 1

                        # Update control finding
                        query = f"""
                            UPDATE control_findings
                            SET {', '.join(set_clauses)}
                            WHERE assessment_id = $1 AND control_id = $2
                        """

                        result = await conn.execute(query, *params)

                        # Check if row was updated
                        if result == "UPDATE 0":
                            # Insert if doesn't exist
                            await conn.execute("""
                                INSERT INTO control_findings (
                                    id, assessment_id, control_id, status,
                                    implementation_status, implementation_narrative,
                                    test_results, findings, recommendations,
                                    risk_level, residual_risk, created_by
                                ) VALUES (
                                    $1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12
                                )
                            """,
                                uuid.uuid4(),
                                assessment_id,
                                control_id,
                                status,
                                update.get('implementation_status'),
                                update.get('implementation_narrative'),
                                update.get('test_results'),
                                update.get('findings'),
                                update.get('recommendations'),
                                update.get('risk_level'),
                                update.get('residual_risk'),
                                updated_by
                            )

                        success_count += 1

                    except Exception as e:
                        logger.error(f"Error updating control {update.get('control_id')}: {e}")
                        errors.append({
                            'control_id': update.get('control_id'),
                            'error': str(e)
                        })
                        failure_count += 1

                # Update assessment progress
                await self._update_assessment_progress(conn, assessment_id)

        return {
            'operation': BulkOperationType.CONTROL_UPDATE.value,
            'total': len(updates),
            'success': success_count,
            'failed': failure_count,
            'errors': errors,
            'status': BulkOperationStatus.COMPLETED.value if failure_count == 0
                     else BulkOperationStatus.PARTIAL.value if success_count > 0
                     else BulkOperationStatus.FAILED.value
        }

    # ========================================================================
    # Bulk Evidence Upload
    # ========================================================================

    async def bulk_upload_evidence_zip(
        self,
        assessment_id: str,
        organization_id: str,
        zip_file: BinaryIO,
        evidence_type: str,
        control_ids: Optional[List[str]],
        uploaded_by: str,
        storage_path: str = "/var/cmmc/evidence"
    ) -> Dict[str, Any]:
        """
        Upload multiple evidence files from a ZIP archive

        Args:
            assessment_id: Assessment ID
            organization_id: Organization ID
            zip_file: ZIP file binary stream
            evidence_type: Type of evidence (Policy, Procedure, etc.)
            control_ids: Optional list of control IDs to link all evidence to
            uploaded_by: User ID performing upload
            storage_path: Base path for evidence storage

        Returns:
            dict: Results with uploaded file count

        Example:
            with open('evidence.zip', 'rb') as f:
                result = await service.bulk_upload_evidence_zip(
                    assessment_id='uuid',
                    organization_id='uuid',
                    zip_file=f,
                    evidence_type='Policy',
                    control_ids=['AC.L2-3.1.1', 'AC.L2-3.1.2'],
                    uploaded_by='user_id'
                )
        """
        import os
        import hashlib

        success_count = 0
        failure_count = 0
        errors = []
        uploaded_files = []

        try:
            with zipfile.ZipFile(zip_file, 'r') as zf:
                # Get list of files (exclude directories and hidden files)
                file_list = [
                    f for f in zf.namelist()
                    if not f.endswith('/') and not os.path.basename(f).startswith('.')
                ]

                logger.info(f"Processing ZIP with {len(file_list)} files")

                async with self.db_pool.acquire() as conn:
                    for file_name in file_list:
                        try:
                            # Read file content
                            file_data = zf.read(file_name)
                            file_size = len(file_data)

                            # Skip empty files
                            if file_size == 0:
                                logger.warning(f"Skipping empty file: {file_name}")
                                continue

                            # Calculate SHA-256 hash
                            file_hash = hashlib.sha256(file_data).hexdigest()

                            # Generate unique filename
                            evidence_id = str(uuid.uuid4())
                            base_name = os.path.basename(file_name)
                            safe_filename = f"{evidence_id}_{base_name}"

                            # Determine file path
                            file_path = os.path.join(
                                storage_path,
                                organization_id,
                                assessment_id,
                                safe_filename
                            )

                            # Create directory if needed
                            os.makedirs(os.path.dirname(file_path), exist_ok=True)

                            # Write file to disk
                            with open(file_path, 'wb') as f:
                                f.write(file_data)

                            # Determine MIME type
                            import mimetypes
                            mime_type, _ = mimetypes.guess_type(base_name)

                            # Insert evidence record
                            await conn.execute("""
                                INSERT INTO evidence (
                                    id, assessment_id, organization_id,
                                    file_name, file_path, file_size_bytes,
                                    file_type, mime_type, file_hash,
                                    title, evidence_type, control_ids,
                                    collection_date, collected_by, created_by, status
                                ) VALUES (
                                    $1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14, $15, $16
                                )
                            """,
                                uuid.UUID(evidence_id),
                                uuid.UUID(assessment_id),
                                uuid.UUID(organization_id),
                                base_name,
                                file_path,
                                file_size,
                                os.path.splitext(base_name)[1][1:] if '.' in base_name else 'unknown',
                                mime_type,
                                file_hash,
                                base_name,  # title = filename
                                evidence_type,
                                control_ids if control_ids else [],
                                datetime.utcnow(),
                                uuid.UUID(uploaded_by),
                                uuid.UUID(uploaded_by),
                                'Active'
                            )

                            uploaded_files.append({
                                'id': evidence_id,
                                'file_name': base_name,
                                'file_size': file_size,
                                'file_path': file_path
                            })

                            success_count += 1

                        except Exception as e:
                            logger.error(f"Error processing file {file_name}: {e}")
                            errors.append({
                                'file_name': file_name,
                                'error': str(e)
                            })
                            failure_count += 1

        except zipfile.BadZipFile as e:
            logger.error(f"Invalid ZIP file: {e}")
            return {
                'operation': BulkOperationType.EVIDENCE_UPLOAD.value,
                'total': 0,
                'success': 0,
                'failed': 1,
                'errors': [{'error': 'Invalid ZIP file'}],
                'status': BulkOperationStatus.FAILED.value
            }

        return {
            'operation': BulkOperationType.EVIDENCE_UPLOAD.value,
            'total': success_count + failure_count,
            'success': success_count,
            'failed': failure_count,
            'errors': errors,
            'uploaded_files': uploaded_files,
            'status': BulkOperationStatus.COMPLETED.value if failure_count == 0
                     else BulkOperationStatus.PARTIAL.value if success_count > 0
                     else BulkOperationStatus.FAILED.value
        }

    # ========================================================================
    # Excel Import/Export
    # ========================================================================

    async def export_findings_to_excel(
        self,
        assessment_id: str
    ) -> io.BytesIO:
        """
        Export control findings to Excel workbook

        Args:
            assessment_id: Assessment ID

        Returns:
            BytesIO: Excel file in memory

        Format:
            - Control ID
            - Domain
            - Title
            - Status
            - Implementation Narrative
            - Findings
            - Recommendations
            - Risk Level
            - Assigned To
        """
        async with self.db_pool.acquire() as conn:
            # Get assessment info
            assessment = await conn.fetchrow("""
                SELECT name, assessment_type, target_level
                FROM assessments
                WHERE id = $1
            """, uuid.UUID(assessment_id))

            # Get control findings
            findings = await conn.fetch("""
                SELECT
                    cf.control_id,
                    c.domain,
                    c.title,
                    cf.status,
                    cf.implementation_status,
                    cf.implementation_narrative,
                    cf.test_results,
                    cf.findings,
                    cf.recommendations,
                    cf.risk_level,
                    cf.residual_risk,
                    u.full_name as assigned_to,
                    cf.examine_completed,
                    cf.interview_completed,
                    cf.test_completed,
                    cf.ai_confidence_score,
                    cf.updated_at
                FROM control_findings cf
                JOIN cmmc_controls c ON cf.control_id = c.id
                LEFT JOIN users u ON cf.assigned_to = u.id
                WHERE cf.assessment_id = $1
                ORDER BY c.domain, cf.control_id
            """, uuid.UUID(assessment_id))

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Control Findings"

        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            "Control ID", "Domain", "Title", "Status", "Implementation Status",
            "Implementation Narrative", "Test Results", "Findings",
            "Recommendations", "Risk Level", "Residual Risk", "Assigned To",
            "Examine", "Interview", "Test", "AI Confidence", "Last Updated"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Data rows
        for row_num, finding in enumerate(findings, 2):
            data = [
                finding['control_id'],
                finding['domain'],
                finding['title'],
                finding['status'],
                finding['implementation_status'],
                finding['implementation_narrative'],
                finding['test_results'],
                finding['findings'],
                finding['recommendations'],
                finding['risk_level'],
                finding['residual_risk'],
                finding['assigned_to'],
                'Yes' if finding['examine_completed'] else 'No',
                'Yes' if finding['interview_completed'] else 'No',
                'Yes' if finding['test_completed'] else 'No',
                f"{finding['ai_confidence_score']:.2f}" if finding['ai_confidence_score'] else '',
                finding['updated_at'].strftime('%Y-%m-%d %H:%M') if finding['updated_at'] else ''
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        # Adjust column widths
        column_widths = [15, 8, 40, 15, 18, 50, 30, 30, 30, 12, 30, 20, 8, 10, 8, 12, 18]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add metadata sheet
        meta_ws = wb.create_sheet("Metadata")
        meta_ws.append(["Assessment Name", assessment['name'] if assessment else ''])
        meta_ws.append(["Assessment Type", assessment['assessment_type'] if assessment else ''])
        meta_ws.append(["Target Level", assessment['target_level'] if assessment else ''])
        meta_ws.append(["Export Date", datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')])
        meta_ws.append(["Total Controls", len(findings)])

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    async def import_findings_from_excel(
        self,
        assessment_id: str,
        excel_file: BinaryIO,
        updated_by: str
    ) -> Dict[str, Any]:
        """
        Import control findings from Excel workbook

        Args:
            assessment_id: Assessment ID
            excel_file: Excel file binary stream
            updated_by: User ID performing import

        Returns:
            dict: Results with import statistics

        Expected Format:
            - Column A: Control ID (required)
            - Column D: Status (required)
            - Column E: Implementation Status
            - Column F: Implementation Narrative
            - Column G: Test Results
            - Column H: Findings
            - Column I: Recommendations
            - Column J: Risk Level
        """
        success_count = 0
        failure_count = 0
        errors = []

        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True)
            ws = wb.active

            # Skip header row
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            updates = []
            for row_num, row in enumerate(rows, 2):
                try:
                    control_id = row[0]  # Column A
                    status = row[3]      # Column D

                    if not control_id or not status:
                        logger.warning(f"Skipping row {row_num}: missing control_id or status")
                        continue

                    update = {
                        'control_id': str(control_id).strip(),
                        'status': str(status).strip(),
                        'implementation_status': row[4] if len(row) > 4 and row[4] else None,
                        'implementation_narrative': row[5] if len(row) > 5 and row[5] else None,
                        'test_results': row[6] if len(row) > 6 and row[6] else None,
                        'findings': row[7] if len(row) > 7 and row[7] else None,
                        'recommendations': row[8] if len(row) > 8 and row[8] else None,
                        'risk_level': row[9] if len(row) > 9 and row[9] else None,
                    }

                    updates.append(update)

                except Exception as e:
                    logger.error(f"Error parsing row {row_num}: {e}")
                    errors.append({
                        'row': row_num,
                        'error': f"Parse error: {str(e)}"
                    })
                    failure_count += 1

            # Use bulk update
            if updates:
                result = await self.bulk_update_control_status(
                    assessment_id, updates, updated_by
                )
                success_count = result['success']
                failure_count += result['failed']
                errors.extend(result['errors'])

        except Exception as e:
            logger.error(f"Error importing Excel: {e}")
            return {
                'operation': BulkOperationType.EXCEL_IMPORT.value,
                'total': 0,
                'success': 0,
                'failed': 1,
                'errors': [{'error': f"Import failed: {str(e)}"}],
                'status': BulkOperationStatus.FAILED.value
            }

        return {
            'operation': BulkOperationType.EXCEL_IMPORT.value,
            'total': success_count + failure_count,
            'success': success_count,
            'failed': failure_count,
            'errors': errors,
            'status': BulkOperationStatus.COMPLETED.value if failure_count == 0
                     else BulkOperationStatus.PARTIAL.value if success_count > 0
                     else BulkOperationStatus.FAILED.value
        }

    # ========================================================================
    # Mass Assignments
    # ========================================================================

    async def mass_assign_controls(
        self,
        assessment_id: str,
        control_ids: List[str],
        assigned_to: str,
        updated_by: str
    ) -> Dict[str, Any]:
        """
        Assign multiple controls to a user

        Args:
            assessment_id: Assessment ID
            control_ids: List of control IDs to assign
            assigned_to: User ID to assign controls to
            updated_by: User ID performing assignment

        Returns:
            dict: Results with assignment statistics
        """
        success_count = 0
        failure_count = 0
        errors = []

        async with self.db_pool.acquire() as conn:
            async with conn.transaction():
                for control_id in control_ids:
                    try:
                        # Update or insert control finding
                        result = await conn.execute("""
                            UPDATE control_findings
                            SET assigned_to = $3, updated_at = NOW(), created_by = $4
                            WHERE assessment_id = $1 AND control_id = $2
                        """, uuid.UUID(assessment_id), control_id, uuid.UUID(assigned_to), uuid.UUID(updated_by))

                        # If no row updated, insert new finding
                        if result == "UPDATE 0":
                            await conn.execute("""
                                INSERT INTO control_findings (
                                    id, assessment_id, control_id, assigned_to, created_by
                                ) VALUES ($1, $2, $3, $4, $5)
                            """,
                                uuid.uuid4(),
                                uuid.UUID(assessment_id),
                                control_id,
                                uuid.UUID(assigned_to),
                                uuid.UUID(updated_by)
                            )

                        success_count += 1

                    except Exception as e:
                        logger.error(f"Error assigning control {control_id}: {e}")
                        errors.append({
                            'control_id': control_id,
                            'error': str(e)
                        })
                        failure_count += 1

        return {
            'operation': BulkOperationType.MASS_ASSIGNMENT.value,
            'total': len(control_ids),
            'success': success_count,
            'failed': failure_count,
            'errors': errors,
            'status': BulkOperationStatus.COMPLETED.value if failure_count == 0
                     else BulkOperationStatus.PARTIAL.value if success_count > 0
                     else BulkOperationStatus.FAILED.value
        }

    # ========================================================================
    # Helper Methods
    # ========================================================================

    async def _update_assessment_progress(
        self,
        conn: asyncpg.Connection,
        assessment_id: str
    ) -> None:
        """
        Update assessment progress counters

        Recalculates:
        - total_controls
        - controls_met
        - controls_not_met
        - controls_partial
        - controls_na
        - completion_percentage
        """
        await conn.execute("""
            UPDATE assessments
            SET
                total_controls = (
                    SELECT COUNT(*) FROM control_findings WHERE assessment_id = $1
                ),
                controls_met = (
                    SELECT COUNT(*) FROM control_findings
                    WHERE assessment_id = $1 AND status = 'Met'
                ),
                controls_not_met = (
                    SELECT COUNT(*) FROM control_findings
                    WHERE assessment_id = $1 AND status = 'Not Met'
                ),
                controls_partial = (
                    SELECT COUNT(*) FROM control_findings
                    WHERE assessment_id = $1 AND status = 'Partially Met'
                ),
                controls_na = (
                    SELECT COUNT(*) FROM control_findings
                    WHERE assessment_id = $1 AND status = 'Not Applicable'
                ),
                completion_percentage = (
                    SELECT CASE
                        WHEN COUNT(*) = 0 THEN 0
                        ELSE ROUND(
                            (COUNT(*) FILTER (WHERE status IN ('Met', 'Not Met', 'Partially Met', 'Not Applicable'))::DECIMAL
                             / COUNT(*)) * 100
                        )
                    END
                    FROM control_findings WHERE assessment_id = $1
                ),
                updated_at = NOW()
            WHERE id = $1
        """, uuid.UUID(assessment_id))

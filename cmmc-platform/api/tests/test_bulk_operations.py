"""
Bulk Operations Tests

Comprehensive tests for bulk processing functionality.
"""

import pytest
import asyncio
import zipfile
import io
import os
from datetime import datetime
from unittest.mock import Mock, AsyncMock, MagicMock, patch
import openpyxl

from services.bulk_service import BulkService, BulkOperationType, BulkOperationStatus


# ============================================================================
# Fixtures
# ============================================================================

@pytest.fixture
def mock_db_pool():
    """Mock database connection pool"""
    pool = AsyncMock()

    # Mock connection
    conn = AsyncMock()
    pool.acquire.return_value.__aenter__.return_value = conn
    pool.acquire.return_value.__aexit__.return_value = None

    # Mock transaction
    transaction = AsyncMock()
    conn.transaction.return_value.__aenter__.return_value = transaction
    conn.transaction.return_value.__aexit__.return_value = None

    return pool


@pytest.fixture
def bulk_service(mock_db_pool):
    """Bulk service instance"""
    return BulkService(mock_db_pool)


@pytest.fixture
def sample_assessment_id():
    """Sample assessment UUID"""
    return "550e8400-e29b-41d4-a716-446655440000"


@pytest.fixture
def sample_user_id():
    """Sample user UUID"""
    return "660e8400-e29b-41d4-a716-446655440001"


# ============================================================================
# Bulk Control Update Tests
# ============================================================================

class TestBulkControlUpdates:
    """Test bulk control update functionality"""

    @pytest.mark.asyncio
    async def test_bulk_update_success(self, bulk_service, sample_assessment_id, sample_user_id, mock_db_pool):
        """Test successful bulk control update"""
        # Setup
        updates = [
            {
                'control_id': 'AC.L2-3.1.1',
                'status': 'Met',
                'implementation_narrative': 'Access control implemented',
                'risk_level': 'Low'
            },
            {
                'control_id': 'AC.L2-3.1.2',
                'status': 'Partially Met',
                'findings': 'Some gaps identified'
            }
        ]

        mock_conn = mock_db_pool.acquire.return_value.__aenter__.return_value
        mock_conn.execute.return_value = "UPDATE 1"

        # Execute
        result = await bulk_service.bulk_update_control_status(
            assessment_id=sample_assessment_id,
            updates=updates,
            updated_by=sample_user_id
        )

        # Verify
        assert result['operation'] == BulkOperationType.CONTROL_UPDATE.value
        assert result['total'] == 2
        assert result['success'] == 2
        assert result['failed'] == 0
        assert result['status'] == BulkOperationStatus.COMPLETED.value
        assert len(result['errors']) == 0

    @pytest.mark.asyncio
    async def test_bulk_update_partial_failure(self, bulk_service, sample_assessment_id, sample_user_id, mock_db_pool):
        """Test bulk update with some failures"""
        # Setup
        updates = [
            {'control_id': 'AC.L2-3.1.1', 'status': 'Met'},
            {'control_id': 'INVALID', 'status': 'Met'},  # Will fail
            {'control_id': 'AC.L2-3.1.3', 'status': 'Not Met'}
        ]

        mock_conn = mock_db_pool.acquire.return_value.__aenter__.return_value

        # First update succeeds, second fails, third succeeds
        mock_conn.execute.side_effect = [
            "UPDATE 1",
            Exception("Invalid control ID"),
            "UPDATE 1",
            "UPDATE 1"  # For assessment progress update
        ]

        # Execute
        result = await bulk_service.bulk_update_control_status(
            assessment_id=sample_assessment_id,
            updates=updates,
            updated_by=sample_user_id
        )

        # Verify
        assert result['total'] == 3
        assert result['success'] == 2
        assert result['failed'] == 1
        assert result['status'] == BulkOperationStatus.PARTIAL.value
        assert len(result['errors']) == 1
        assert result['errors'][0]['control_id'] == 'INVALID'

    @pytest.mark.asyncio
    async def test_bulk_update_missing_fields(self, bulk_service, sample_assessment_id, sample_user_id):
        """Test bulk update with missing required fields"""
        # Setup
        updates = [
            {'control_id': 'AC.L2-3.1.1'},  # Missing status
            {'status': 'Met'}  # Missing control_id
        ]

        # Execute
        result = await bulk_service.bulk_update_control_status(
            assessment_id=sample_assessment_id,
            updates=updates,
            updated_by=sample_user_id
        )

        # Verify
        assert result['total'] == 2
        assert result['success'] == 0
        assert result['failed'] == 2
        assert result['status'] == BulkOperationStatus.FAILED.value


# ============================================================================
# Bulk Evidence Upload Tests
# ============================================================================

class TestBulkEvidenceUpload:
    """Test bulk evidence upload functionality"""

    def create_test_zip(self, files: dict) -> io.BytesIO:
        """
        Create test ZIP file

        Args:
            files: Dict of filename: content

        Returns:
            BytesIO: ZIP file in memory
        """
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for filename, content in files.items():
                zf.writestr(filename, content)
        zip_buffer.seek(0)
        return zip_buffer

    @pytest.mark.asyncio
    async def test_bulk_evidence_upload_success(
        self,
        bulk_service,
        sample_assessment_id,
        sample_user_id,
        mock_db_pool,
        tmp_path
    ):
        """Test successful bulk evidence upload"""
        # Setup
        test_files = {
            'policy.pdf': b'PDF content here',
            'procedure.docx': b'DOCX content here',
            'screenshot.png': b'PNG content here'
        }

        zip_file = self.create_test_zip(test_files)

        mock_conn = mock_db_pool.acquire.return_value.__aenter__.return_value
        mock_conn.execute.return_value = None

        # Use temp directory for storage
        storage_path = str(tmp_path)

        # Execute
        result = await bulk_service.bulk_upload_evidence_zip(
            assessment_id=sample_assessment_id,
            organization_id='org-123',
            zip_file=zip_file,
            evidence_type='Policy',
            control_ids=['AC.L2-3.1.1'],
            uploaded_by=sample_user_id,
            storage_path=storage_path
        )

        # Verify
        assert result['operation'] == BulkOperationType.EVIDENCE_UPLOAD.value
        assert result['total'] == 3
        assert result['success'] == 3
        assert result['failed'] == 0
        assert result['status'] == BulkOperationStatus.COMPLETED.value
        assert len(result['uploaded_files']) == 3

        # Verify files were written
        uploaded_count = 0
        for file_info in result['uploaded_files']:
            if os.path.exists(file_info['file_path']):
                uploaded_count += 1
        assert uploaded_count == 3

    @pytest.mark.asyncio
    async def test_bulk_evidence_upload_invalid_zip(
        self,
        bulk_service,
        sample_assessment_id,
        sample_user_id
    ):
        """Test bulk upload with invalid ZIP file"""
        # Setup - not a valid ZIP
        invalid_zip = io.BytesIO(b'This is not a ZIP file')

        # Execute
        result = await bulk_service.bulk_upload_evidence_zip(
            assessment_id=sample_assessment_id,
            organization_id='org-123',
            zip_file=invalid_zip,
            evidence_type='Policy',
            control_ids=None,
            uploaded_by=sample_user_id
        )

        # Verify
        assert result['status'] == BulkOperationStatus.FAILED.value
        assert result['total'] == 0
        assert 'Invalid ZIP file' in result['errors'][0]['error']

    @pytest.mark.asyncio
    async def test_bulk_evidence_upload_empty_files_skipped(
        self,
        bulk_service,
        sample_assessment_id,
        sample_user_id,
        mock_db_pool,
        tmp_path
    ):
        """Test that empty files are skipped"""
        # Setup
        test_files = {
            'valid.pdf': b'PDF content',
            'empty.txt': b'',  # Empty file
            'valid2.docx': b'DOCX content'
        }

        zip_file = self.create_test_zip(test_files)

        mock_conn = mock_db_pool.acquire.return_value.__aenter__.return_value
        mock_conn.execute.return_value = None

        # Execute
        result = await bulk_service.bulk_upload_evidence_zip(
            assessment_id=sample_assessment_id,
            organization_id='org-123',
            zip_file=zip_file,
            evidence_type='Policy',
            control_ids=None,
            uploaded_by=sample_user_id,
            storage_path=str(tmp_path)
        )

        # Verify - empty file should be skipped
        assert result['success'] == 2  # Only 2 valid files
        assert len(result['uploaded_files']) == 2


# ============================================================================
# Excel Import/Export Tests
# ============================================================================

class TestExcelOperations:
    """Test Excel import/export functionality"""

    @pytest.mark.asyncio
    async def test_export_findings_to_excel(
        self,
        bulk_service,
        sample_assessment_id,
        mock_db_pool
    ):
        """Test Excel export"""
        # Setup
        mock_conn = mock_db_pool.acquire.return_value.__aenter__.return_value

        # Mock assessment
        mock_conn.fetchrow.return_value = {
            'name': 'Q1 2024 Assessment',
            'assessment_type': 'CMMC_L2',
            'target_level': 2
        }

        # Mock findings
        mock_conn.fetch.return_value = [
            {
                'control_id': 'AC.L2-3.1.1',
                'domain': 'AC',
                'title': 'Limit System Access',
                'status': 'Met',
                'implementation_status': 'Implemented',
                'implementation_narrative': 'Access control implemented',
                'test_results': 'Passed',
                'findings': None,
                'recommendations': None,
                'risk_level': 'Low',
                'residual_risk': None,
                'assigned_to': 'John Doe',
                'examine_completed': True,
                'interview_completed': True,
                'test_completed': False,
                'ai_confidence_score': 0.95,
                'updated_at': datetime.utcnow()
            }
        ]

        # Execute
        excel_file = await bulk_service.export_findings_to_excel(sample_assessment_id)

        # Verify
        assert isinstance(excel_file, io.BytesIO)

        # Load and verify Excel content
        excel_file.seek(0)
        wb = openpyxl.load_workbook(excel_file)

        # Check sheets
        assert 'Control Findings' in wb.sheetnames
        assert 'Metadata' in wb.sheetnames

        # Check data
        ws = wb['Control Findings']
        assert ws['A1'].value == 'Control ID'
        assert ws['A2'].value == 'AC.L2-3.1.1'
        assert ws['D2'].value == 'Met'

    @pytest.mark.asyncio
    async def test_import_findings_from_excel(
        self,
        bulk_service,
        sample_assessment_id,
        sample_user_id,
        mock_db_pool
    ):
        """Test Excel import"""
        # Setup - Create test Excel file
        wb = openpyxl.Workbook()
        ws = wb.active

        # Headers
        ws.append([
            'Control ID', 'Domain', 'Title', 'Status', 'Implementation Status',
            'Implementation Narrative', 'Test Results', 'Findings',
            'Recommendations', 'Risk Level'
        ])

        # Data
        ws.append([
            'AC.L2-3.1.1', 'AC', 'Limit System Access', 'Met', 'Implemented',
            'Access control implemented', 'Passed', None, None, 'Low'
        ])
        ws.append([
            'AC.L2-3.1.2', 'AC', 'Limit Transactions', 'Partially Met', 'Partially Implemented',
            'Some gaps', None, 'Gaps in enforcement', 'Implement additional controls', 'Moderate'
        ])

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Mock database
        mock_conn = mock_db_pool.acquire.return_value.__aenter__.return_value
        mock_conn.execute.return_value = "UPDATE 1"

        # Execute
        result = await bulk_service.import_findings_from_excel(
            assessment_id=sample_assessment_id,
            excel_file=excel_buffer,
            updated_by=sample_user_id
        )

        # Verify
        assert result['operation'] == BulkOperationType.EXCEL_IMPORT.value
        assert result['success'] == 2
        assert result['failed'] == 0
        assert result['status'] == BulkOperationStatus.COMPLETED.value


# ============================================================================
# Mass Assignment Tests
# ============================================================================

class TestMassAssignment:
    """Test mass assignment functionality"""

    @pytest.mark.asyncio
    async def test_mass_assign_controls_success(
        self,
        bulk_service,
        sample_assessment_id,
        sample_user_id,
        mock_db_pool
    ):
        """Test successful mass assignment"""
        # Setup
        control_ids = ['AC.L2-3.1.1', 'AC.L2-3.1.2', 'AC.L2-3.1.3']
        assigned_to = 'assessor-uuid'

        mock_conn = mock_db_pool.acquire.return_value.__aenter__.return_value
        mock_conn.execute.return_value = "UPDATE 1"

        # Execute
        result = await bulk_service.mass_assign_controls(
            assessment_id=sample_assessment_id,
            control_ids=control_ids,
            assigned_to=assigned_to,
            updated_by=sample_user_id
        )

        # Verify
        assert result['operation'] == BulkOperationType.MASS_ASSIGNMENT.value
        assert result['total'] == 3
        assert result['success'] == 3
        assert result['failed'] == 0
        assert result['status'] == BulkOperationStatus.COMPLETED.value

    @pytest.mark.asyncio
    async def test_mass_assign_with_failures(
        self,
        bulk_service,
        sample_assessment_id,
        sample_user_id,
        mock_db_pool
    ):
        """Test mass assignment with some failures"""
        # Setup
        control_ids = ['AC.L2-3.1.1', 'INVALID', 'AC.L2-3.1.3']
        assigned_to = 'assessor-uuid'

        mock_conn = mock_db_pool.acquire.return_value.__aenter__.return_value

        # First succeeds, second fails, third succeeds
        mock_conn.execute.side_effect = [
            "UPDATE 1",
            Exception("Invalid control"),
            "UPDATE 1"
        ]

        # Execute
        result = await bulk_service.mass_assign_controls(
            assessment_id=sample_assessment_id,
            control_ids=control_ids,
            assigned_to=assigned_to,
            updated_by=sample_user_id
        )

        # Verify
        assert result['total'] == 3
        assert result['success'] == 2
        assert result['failed'] == 1
        assert result['status'] == BulkOperationStatus.PARTIAL.value


# ============================================================================
# Performance Tests
# ============================================================================

class TestBulkPerformance:
    """Test bulk operation performance"""

    @pytest.mark.asyncio
    @pytest.mark.slow
    async def test_bulk_update_large_batch(
        self,
        bulk_service,
        sample_assessment_id,
        sample_user_id,
        mock_db_pool
    ):
        """Test performance with large batch (110 controls)"""
        import time

        # Setup
        updates = [
            {
                'control_id': f'TEST.L2-{i}.{j}',
                'status': 'Met',
                'implementation_narrative': f'Control {i}.{j} implemented'
            }
            for i in range(1, 15)
            for j in range(1, 8)
        ]  # 98 updates

        mock_conn = mock_db_pool.acquire.return_value.__aenter__.return_value
        mock_conn.execute.return_value = "UPDATE 1"

        # Execute and time
        start_time = time.time()

        result = await bulk_service.bulk_update_control_status(
            assessment_id=sample_assessment_id,
            updates=updates,
            updated_by=sample_user_id
        )

        elapsed_time = time.time() - start_time

        # Verify
        assert result['success'] == len(updates)
        assert result['failed'] == 0

        # Should complete quickly (< 5 seconds for mock)
        assert elapsed_time < 5.0

        print(f"Bulk update of {len(updates)} controls completed in {elapsed_time:.2f}s")


if __name__ == '__main__':
    pytest.main([__file__, '-v', '--asyncio-mode=auto'])
